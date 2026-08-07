"""Firma tampon seçim Excel'i → `catalog_data/buffers/firma_excel_buffers.json`

Kaynak: `Tampon Seçimi - Yeni Type S Tipi.xlsx` (tek sayfa: "Tampon Seçimleri")

Excel iki şey taşır:

1. **Ürün listesi** (AE16:BJ19) — 32 sütun × 3 satır: strok [mm], enerji
   kapasitesi [kJ] ve çarpışma kapasitesi [kN]. Bunların 21'i SIBRE SP
   hidrolik tamponudur ve `sibre_sp_hydraulic.json` ile BİREBİR uyuşur
   (betik bu karşılaştırmayı çalıştırır ve sonucu meta'ya yazar). Kalan 11'i
   "Type NN HD" kodlu ürünlerdir; ELİMİZDEKİ ÜRETİCİ KATALOGLARINDA KARŞILIĞI
   YOKTUR. Bu dosyaya YALNIZ o 11 satır yazılır.

2. **Seçim yöntemi** (B3:AA11 formülleri) — dört vinç için işlenmiş örnek.
   Adımlar meta.selection_steps'te formül, birim ve gerekçesiyle durur;
   dört örnek de meta.worked_examples'ta ARA DEĞERLERİYLE saklanır ki sonraki
   fazda kodlanan hesap bunlara karşı sınanabilsin.

Excel `data_only=False` (formüller) ve `data_only=True` (son hesaplanmış
değerler) olarak İKİ KEZ açılır; ikisi de gerekir.
"""

from __future__ import annotations

import json
import re

import openpyxl

from buffers_common import OUT_DIR, XLSX_FIRMA, clean, write_json

SHEET = "Tampon Seçimleri"
CATALOG_ROW = 16   # ürün adı satırı
STROKE_ROW = 17
ENERGY_ROW = 18
FORCE_ROW = 19
CATALOG_COL_FROM = 31  # AE
CATALOG_COL_TO = 62    # BJ
DATA_ROWS = range(4, 12)  # 4..11 — dört vincin araba + köprü satırları

SP_RE = re.compile(r"^Type SP (\d+) - (\d+)$")


def _sheets() -> tuple:
    wf = openpyxl.load_workbook(XLSX_FIRMA, data_only=False)[SHEET]
    wv = openpyxl.load_workbook(XLSX_FIRMA, data_only=True)[SHEET]
    return wf, wv


def _catalog(wv) -> list[tuple[str, float, float, float]]:
    rows = []
    for c in range(CATALOG_COL_FROM, CATALOG_COL_TO + 1):
        name = wv.cell(CATALOG_ROW, c).value
        if not name:
            continue
        rows.append((
            str(name).strip(),
            wv.cell(STROKE_ROW, c).value,
            wv.cell(ENERGY_ROW, c).value,
            wv.cell(FORCE_ROW, c).value,
        ))
    return rows


def _verify_sp(rows) -> dict:
    """Excel'deki SP satırlarını SIBRE kataloğu çıkarımıyla karşılaştırır."""
    try:
        with open(f"{OUT_DIR}/sibre_sp_hydraulic.json", encoding="utf-8") as fh:
            sibre = json.load(fh)["items"]
    except FileNotFoundError:
        return {"status": "sibre_sp_hydraulic.json yok — karşılaştırma yapılmadı"}

    by_key: dict[tuple[str, int], dict] = {}
    for it in sibre:
        by_key.setdefault((it["series"], it["stroke_mm"]), it)

    matched, mismatched, missing = 0, [], []
    for name, stroke, energy, force in rows:
        m = SP_RE.match(name)
        if not m:
            continue
        it = by_key.get((f"SP {m.group(1)}", int(stroke)))
        if it is None:
            missing.append(name)
        elif it["energy_capacity_kj"] == energy and it["max_end_force_kn"] == force:
            matched += 1
        else:
            mismatched.append({
                "model": name,
                "excel": {"energy_kj": energy, "force_kn": force},
                "katalog": {
                    "energy_kj": it["energy_capacity_kj"],
                    "force_kn": it["max_end_force_kn"],
                },
            })
    return {
        "checked": matched + len(mismatched) + len(missing),
        "matched": matched,
        "mismatched": mismatched,
        "missing_in_catalog": missing,
    }


def _worked_examples(wf, wv) -> list[dict]:
    out: list[dict] = []
    name = None
    for r in DATA_ROWS:
        if wv.cell(r, 3).value:  # C sütunu = vinç adı
            name = str(wv.cell(r, 3).value).strip()
        if wv.cell(r, 5).value is None:  # E sütunu = yük grubu
            continue
        out.append({
            "crane": name,
            "row": r,
            "group": wv.cell(r, 5).value,
            "inputs": {
                "mass_kg": wv.cell(r, 4).value,
                "speed_m_min": wv.cell(r, 6).value,
                "min_trolley_approach_m": wv.cell(r, 7).value
                if isinstance(wv.cell(r, 7).value, (int, float)) else None,
                "span_m": wv.cell(r, 8).value
                if isinstance(wv.cell(r, 8).value, (int, float)) else None,
                "motor_power_kw": wv.cell(r, 11).value,
                "motor_count": wv.cell(r, 12).value,
                "motor_speed_rpm": wv.cell(r, 13).value,
                "gear_ratio": wv.cell(r, 14).value,
            },
            "selected_buffer": wv.cell(r, 19).value,
            "intermediate": {
                "impact_mass_kg": wv.cell(r, 9).value,
                "impact_energy_kj": wv.cell(r, 10).value,
                "drive_force_per_buffer_n": wv.cell(r, 15).value,
                "drive_energy_kj": wv.cell(r, 16).value,
                "total_energy_kj": wv.cell(r, 17).value,
                "buffer_load_kn": wv.cell(r, 18).value,
                "buffer_energy_kj": wv.cell(r, 20).value,
                "buffer_stroke_mm": wv.cell(r, 21).value,
                "buffer_force_kn": wv.cell(r, 22).value,
                "safety_factor": wv.cell(r, 25).value,
                "deceleration_m_s2": wv.cell(r, 26).value,
            },
            "checks": {
                "force_check": wv.cell(r, 23).value,
                "energy_check": wv.cell(r, 24).value,
                "deceleration_check": wv.cell(r, 27).value,
            },
        })
    return out


SELECTION_STEPS = [
    {
        "step": 1,
        "cell": "I",
        "name": "Çarpışma kütlesi",
        "unit": "kg",
        "formula_trolley": "I = D_araba",
        "formula_bridge": "I = D_köprü / 2 + D_araba × (H − G) / H",
        "note": "Araba çarpmasında arabanın kendi ağırlığı. Köprü çarpmasında "
                "köprü ağırlığının yarısı (iki tampon paralel) + arabanın o "
                "tarafa düşen payı; H açıklık, G arabanın minimum yanaşma "
                "mesafesi. YÜK (kanca kütlesi) DAHİL DEĞİLDİR.",
    },
    {
        "step": 2,
        "cell": "J",
        "name": "Çarpışma enerjisi",
        "unit": "kJ",
        "formula": "J = (I / 1000) × (F / 60 × 0,7)² × 0,5",
        "note": "E = ½ m v². Kütle tona çevrilir (t·(m/s)² = kJ). Çarpma hızı "
                "anma yürüyüş hızının %70'idir (F m/dak → F/60 m/s).",
    },
    {
        "step": 3,
        "cell": "O",
        "name": "Tampon başına toplam yürütme yükü",
        "unit": "N (Excel'in kabulü)",
        "formula": "O = 9550 × K × L / M × N / 2",
        "note": "9550·P[kW]/n[d/dak] = motor momenti [Nm]; × motor adedi (L) × "
                "redüktör çevrim oranı (N) = tahrik çıkış momenti; / 2 = tampon "
                "başına. DİKKAT — BOYUT TUTARSIZ: sonuç Nm'dir ama sonraki "
                "adımlarda KUVVET (N) gibi kullanılır; yani teker yarıçapı "
                "örtük olarak 1 m alınmıştır. Sonraki fazda bu, tahrik "
                "kuvveti = çıkış momenti / teker yarıçapı olarak "
                "DÜZELTİLMELİDİR.",
    },
    {
        "step": 4,
        "cell": "P",
        "name": "Yürütmeden gelen enerji",
        "unit": "kJ",
        "formula": "P = O × U / 1 000 000",
        "note": "Tahrik kuvveti [N] × tampon stroku [mm] = [mJ]; /1e6 = kJ. "
                "Motor çarpma boyunca tahrik etmeye devam eder kabulü. "
                "U seçilen tamponun stroku olduğundan hesap DÖNGÜSELDİR: önce "
                "tampon seçilir, sonra doğrulanır.",
    },
    {
        "step": 5,
        "cell": "Q",
        "name": "Toplam sönümlenmesi gereken enerji",
        "unit": "kJ",
        "formula": "Q = P + J",
    },
    {
        "step": 6,
        "cell": "R",
        "name": "Tampon yükü",
        "unit": "kN",
        "formula": "R = Q / (U/1000 × 0,8) + O / 1000",
        "note": "Sönümleme verimi 0,8 alınmıştır. DİKKAT — SIBRE kataloğu (s.18) "
                "kendi son kuvvetlerini 0,85 verimle basar: 'The specified final "
                "forces are already applied with a damping efficiency of 0.85'. "
                "Firma Excel'i 0,8 ile daha muhafazakâr davranır (aynı enerjide "
                "%6 daha yüksek kuvvet). Tahrik kuvveti kuvvete de eklenir.",
    },
    {
        "step": 7,
        "cell": "S",
        "name": "Tampon seçimi",
        "unit": "—",
        "formula": "mühendis listeden seçer",
        "note": "T/U/V, seçilen ada göre AE16:BJ19 tablosundan "
                "INDEX+MATCH ile okunur: T = enerji kapasitesi [kJ], "
                "U = strok [mm], V = çarpışma kapasitesi [kN].",
    },
    {
        "step": 8,
        "cell": "W",
        "name": "Kuvvet kontrolü",
        "unit": "—",
        "formula": "V ≥ R → OK",
    },
    {
        "step": 9,
        "cell": "X",
        "name": "Enerji kontrolü",
        "unit": "—",
        "formula": "T ≥ Q → OK",
    },
    {
        "step": 10,
        "cell": "Y",
        "name": "Sağlanan emniyet katsayısı",
        "unit": "—",
        "formula": "Y = T / Q",
        "note": "Bilgi amaçlı; Excel'de alt sınır kontrolü YOKTUR. "
                "Örneklerde 1,70 ile 18,1 arasında değerler geçer.",
    },
    {
        "step": 11,
        "cell": "Z",
        "name": "Yavaşlama ivmesi",
        "unit": "m/s²",
        "formula": "Z = (F/60) / ((U/1000) / (F/60 / 2)) = v² / (2 s)",
        "note": "Sabit yavaşlama kabulü. DİKKAT: burada ANMA hızı (v) "
                "kullanılır, adım 2'deki 0,7·v DEĞİL.",
    },
    {
        "step": 12,
        "cell": "AA",
        "name": "İvme kontrolü",
        "unit": "—",
        "formula": "Z ≤ 5 → OK",
        "note": "Firma tasarım kabulü (kaynak standart Excel'de belirtilmemiş).",
    },
]


def build() -> tuple[str, int]:
    wf, wv = _sheets()
    rows = _catalog(wv)

    items = []
    for name, stroke, energy, force in rows:
        if SP_RE.match(name):
            continue
        items.append({
            "model": name,
            "type": "bilinmiyor",
            "stroke_mm": clean(stroke),
            "energy_capacity_kj": clean(energy),
            "max_force_kn": clean(force),
            "source": "firma Excel'i",
        })
    items.sort(key=lambda r: (r["model"]))

    path = write_json(
        "firma_excel_buffers.json",
        {
            "brand": None,
            "equipment_type": "buffer",
            "source_xlsx": "Tampon Seçimi - Yeni Type S Tipi.xlsx",
            "source_sheet": SHEET,
            "source_range": "AE16:BJ19",
            "extraction_date": "2026-08-06",
            "sibre_sp_verification": _verify_sp(rows),
            "selection_steps": SELECTION_STEPS,
            "worked_examples": _worked_examples(wf, wv),
            "notes": (
                "kaynak: firma Excel'i, üretici kataloğuyla teyit EDİLMEDİ. "
                "Bu dosyadaki 'Type NN HD' modellerinin elimizdeki üretici "
                "kataloglarında (SIBRE SP, Conductix 0170/0180) KARŞILIĞI "
                "YOKTUR; marka, malzeme, ağırlık ve montaj ölçüleri BİLİNMİYOR "
                "— uydurulmamış, alanlar açılmamıştır. Dosya adı 'Yeni Type S "
                "Tipi' olduğundan bunların bir 'Type S' hidrolik tampon "
                "serisi olması muhtemeldir, ancak DOĞRULANMAMIŞTIR; katalog "
                "elde edilene kadar seçici listesine ALINMAMALIDIR. "
                "Aynı Excel'deki 21 SIBRE SP satırının tamamı "
                "sibre_sp_hydraulic.json ile birebir uyuşur "
                "(meta.sibre_sp_verification) — bu, SIBRE çıkarımının bağımsız "
                "doğrulamasıdır. "
                "Excel'in hesap adımları meta.selection_steps'te, dört işlenmiş "
                "örnek meta.worked_examples'ta durur."
            ),
        },
        ["model", "type", "stroke_mm", "energy_capacity_kj", "max_force_kn",
         "source"],
        items,
    )
    return path, len(items)


if __name__ == "__main__":
    path, n = build()
    print(f"yazıldı: {path}  ({n} satır)")
    with open(path, encoding="utf-8") as fh:
        meta = json.load(fh)["meta"]
    v = meta["sibre_sp_verification"]
    print(f"SIBRE SP karşılaştırması: {v['matched']}/{v['checked']} birebir, "
          f"{len(v['mismatched'])} farklı, "
          f"{len(v['missing_in_catalog'])} katalogda yok "
          f"({v['missing_in_catalog']})")
